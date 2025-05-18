from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.urls import reverse
from decimal import Decimal
import json
from django.db.models import Sum, F,Q
from datetime import datetime
from .forms import CustomUserCreationForm, TrainingClassForm
from .models import Student, Payment, PendingPayment, SalesPerson, TrainingClass
from django.contrib import messages
from django.core.mail import send_mail
# Sales Home View
@login_required
def sales_home(request):
    return render(request, "sales/home.html")

# Admin Dashboard

@login_required
def admin_dashboard(request):
    if not request.user.role == 'admin':
        return redirect('home')

    # Get filter parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Convert dates to datetime objects
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Fetch data with filters
    total_students = Student.objects.filter(is_enrolled=True).count()  # Only count fully enrolled students
    student_progress = []
    for student in Student.objects.filter(is_enrolled=True):  # Only include fully enrolled students
        total_paid = Payment.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total'] or 0
        progress = (total_paid / student.total_fees) * 100 if student.total_fees > 0 else 0
        student_progress.append({
            "name": student.name,
            "training_class": student.training_class.name,
            "total_fees": student.total_fees,
            "total_paid": total_paid,
            "progress": round(progress, 2),  # Round to 2 decimal places
        })

    # Apply date range filter to total_collected
    payments = Payment.objects.all()

    if start_date and end_date:
        student_data = Student.objects.all()
        payments = payments.filter(payment_date__range=[start_date, end_date])
        total_students = student_data.filter(joined_date__range=[start_date, end_date]).filter(is_enrolled=True).count()
        total_students_data = student_data.filter(joined_date__range=[start_date, end_date]).filter(is_enrolled=True)
        student_progress = []
        for student in total_students_data:  # Only include fully enrolled students
            total_paid = Payment.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total'] or 0
            progress = (total_paid / student.total_fees) * 100 if student.total_fees > 0 else 0
            student_progress.append({
                "name": student.name,
                "training_class": student.training_class.name,
                "total_fees": student.total_fees,
                "total_paid": total_paid,
                "progress": round(progress, 2),  # Round to 2 decimal places
            })

    total_collected = payments.aggregate(total=Sum('amount_paid'))['total'] or 0

    # Filter pending payments based on the payment date of related payments
    pending_payments = PendingPayment.objects.all()
    if start_date and end_date:
        pending_payments = pending_payments.filter(student__payment__payment_date__range=[start_date, end_date]).distinct()

    total_pending = pending_payments.aggregate(total=Sum('due_amount'))['total'] or 0

    sales_data = SalesPerson.objects.all().annotate(
        total_collection=Sum(
            'payment__amount_paid',
            filter=Q(payment__payment_date__range=[start_date, end_date]) if start_date and end_date else Q()
        )
    ).filter(total_collection__gt=0)

    # Prepare data for the bar chart
    salespersons = [sales.user.get_full_name() or sales.user.username for sales in sales_data if sales.user.role == 'sales']
    collections = [float(sales.total_collection or 0) for sales in sales_data]

    # Filter students based on the date range for the pie chart
    students = Student.objects.filter(is_enrolled=True)  # Only include fully enrolled students
    if start_date and end_date:
        students = students.filter(payment__payment_date__range=[start_date, end_date])

    # Calculate payment status counts for the pie chart
    paid_count = students.filter(payment_status='paid').count()
    pending_count = students.filter(payment_status='pending').count()
    not_updated_count = students.filter(payment_status='not_updated').count()

    # Fetch pending payments
    pending_payments = PendingPayment.objects.filter(due_amount__gt=0)

    # Fetch student progress data (only for fully enrolled students)


    context = {
        'total_students': total_students,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'sales_data': sales_data,
        'salespersons': json.dumps(salespersons),  # Convert to JSON for Chart.js
        'collections': json.dumps(collections),    # Convert to JSON for Chart.js
        'paid_count': paid_count,
        'pending_count': pending_count,
        'not_updated_count': not_updated_count,
        'pending_payments': pending_payments,      # Pass pending payments to the template
        'student_progress': student_progress,      # Pass student progress to the template
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
    }
    return render(request, 'admin_dashboard.html', context)
# Custom Login View
class CustomLoginView(LoginView):
    def get_success_url(self):
        user = self.request.user
        if user.role == "sales":
            return reverse("sales_dashboard")
        elif user.role == "admin":
            return reverse("admin_dashboard")
        return reverse("home")

# Salesperson Permission Decorator
def sales_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and getattr(request.user, "role", None) == "sales":
            return view_func(request, *args, **kwargs)
        return redirect("login")
    return wrapper

# Sales Dashboard
@login_required
@sales_required
def sales_dashboard(request):
    try:
        # Get the salesperson associated with the logged-in user
        salesperson = SalesPerson.objects.get(user=request.user)
    except SalesPerson.DoesNotExist:
        # If no salesperson exists, create one (optional, depending on your logic)
        salesperson = SalesPerson.objects.create(user=request.user)

    # Fetch data for the dashboard
    students = Student.objects.filter(salesperson=salesperson)
    total_collected = Payment.objects.filter(collected_by=salesperson).aggregate(total=Sum("amount_paid"))["total"] or 0
    pending_payments = PendingPayment.objects.filter(student__salesperson=salesperson, due_amount__gt=0)
    fully_paid_students = students.filter(payment_status='paid')

    context = {
        "students": students,
        "total_collected": total_collected,
        "pending_payments": pending_payments,
        "fully_paid_students": fully_paid_students,
    }
    return render(request, "sales_dashboard.html", context)

# Enroll a Student
@login_required
def enroll_student(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        training_class_id = request.POST.get("training_class")
        total_fees = request.POST.get("total_fees")

        # Check if a student with the same email already exists
        if Student.objects.filter(email=email).exists():
            messages.error(request, "A student with this email is already enrolled.")
            return redirect("enroll_student")

        try:
            # Fetch the training class
            training_class = TrainingClass.objects.get(id=training_class_id)

            # Fetch the salesperson associated with the logged-in user
            salesperson = SalesPerson.objects.get(user=request.user)

            # Create the student
            Student.objects.create(
                name=name,
                email=email,
                phone=phone,
                training_class=training_class,
                total_fees=total_fees,
                payment_status='pending',
                salesperson=salesperson,
                is_enrolled=True  # Mark as fully enrolled
            )
            subject = "Enrollment Confirmation"
            message = f"Hello {name},\n\nYou have been successfully enrolled for {training_class.name} course. Your total fee is {total_fees}.\ntraining details:\nstart_date - {training_class.start_date}\n end_date- {training_class.end_date} \nBest regards,\nTraining Team"

            from_email = "jsreddychanda123@gmail.com"  # Replace with your email
            recipient_list = [email]

            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            messages.success(request, f"Student {name} enrolled successfully! A confirmation email has been sent.")

            #messages.success(request, f"Student {name} enrolled successfully!")
            return redirect("sales_dashboard")

        except TrainingClass.DoesNotExist:
            messages.error(request, "Invalid training class selected.")
            return redirect("enroll_student")
        except SalesPerson.DoesNotExist:
            messages.error(request, "You are not authorized to enroll students.")
            return redirect("enroll_student")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect("enroll_student")

    # Fetch all training classes for the form dropdown
    classes = TrainingClass.objects.all()
    context = {
        "classes": classes,
    }
    return render(request, "sales/enroll_student.html", context)

@login_required
def record_payment(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    salesperson = get_object_or_404(SalesPerson, user=request.user)
    pending_payment = PendingPayment.objects.filter(student=student).first()

    # Calculate total paid and remaining fees
    total_paid = Payment.objects.filter(student=student).aggregate(total=Sum('amount_paid'))['total'] or 0
    remaining_fees = student.total_fees - total_paid

    # Prevent payments if the student has already paid their total fees
    if remaining_fees <= 0:
        messages.error(request, f"{student.name} has already paid their total fees. No further payments are allowed.")
        return redirect("sales_dashboard")

    if request.method == "POST":
        try:
            amount_paid = Decimal(request.POST.get("amount_paid"))
            payment_status = request.POST.get("payment_status")  # 'full' or 'partial'

            # Validate the amount paid
            if amount_paid <= 0:
                messages.error(request, "Amount paid must be greater than zero.")
                return redirect("record_payment", student_id=student.id)
            if amount_paid > remaining_fees:
                messages.error(request, f"Amount paid cannot exceed the remaining fees of ₹{remaining_fees}.")
                return redirect("record_payment", student_id=student.id)

            # Save Payment
            Payment.objects.create(student=student, collected_by=salesperson, amount_paid=amount_paid)

            if payment_status == "full" or amount_paid >= remaining_fees:
                # Mark as fully paid
                student.payment_status = 'paid'
                student.save()
                # Clear any pending payment for the student
                PendingPayment.objects.filter(student=student).delete()
                messages.success(request, f"Payment of ₹{amount_paid} recorded. {student.name} is now fully paid.")
            else:
                # Handle partial payment
                if pending_payment:
                    pending_payment.due_amount -= amount_paid
                    if pending_payment.due_amount <= 0:
                        # If due amount is fully paid, delete the pending payment
                        pending_payment.delete()
                        student.payment_status = 'paid'
                        messages.success(request, f"Payment of ₹{amount_paid} recorded. {student.name} is now fully paid.")
                    else:
                        # Update the pending payment
                        pending_payment.save()
                        student.payment_status = 'pending'
                        messages.success(request, f"Payment of ₹{amount_paid} recorded. {student.name} has a pending amount of ₹{pending_payment.due_amount}.")
                else:
                    # Create a new pending payment if it doesn't exist
                    PendingPayment.objects.create(student=student, due_amount=remaining_fees - amount_paid)
                    student.payment_status = 'pending'
                    messages.success(request, f"Payment of ₹{amount_paid} recorded. {student.name} has a pending amount of ₹{remaining_fees - amount_paid}.")

                # Save the updated payment status
                student.save()

            return redirect("sales_dashboard")

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect("record_payment", student_id=student.id)

    context = {
        "student": student,
        "remaining_fees": remaining_fees,
    }
    return render(request, "sales/record_payment.html", context)

# View Pending Payments
@login_required
@sales_required
def view_pending_payments(request):
    salesperson = get_object_or_404(SalesPerson, user=request.user)
    pending_payments = PendingPayment.objects.filter(student__salesperson=salesperson, due_amount__gt=0)
    return render(request, "sales/pending_payments.html", {"pending_payments": pending_payments})

# Create Pending Payment
@login_required
def create_pending_payment(request, student_id):
    student = Student.objects.get(id=student_id)

    # Check if the student already has a pending payment
    if PendingPayment.objects.filter(student=student).exists():
        messages.error(request, f"{student.name} already has a pending payment.")
        return redirect('sales_dashboard')  # Redirect to the dashboard or another page

    # If no pending payment exists, create one
    PendingPayment.objects.create(student=student, due_amount=1000)  # Adjust due_amount as needed
    messages.success(request, f"Pending payment created for {student.name}.")
    return redirect('sales_dashboard')

# Add User
@login_required
def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'User added successfully!')
            return redirect('admin_dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'sales/add_user.html', {'form': form})

# Add Training Class
@login_required
def add_class(request):
    if request.method == 'POST':
        form = TrainingClassForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Training class added successfully!')
            return redirect('admin_dashboard')
    else:
        form = TrainingClassForm()
    return render(request, 'sales/add_class.html', {'form': form})

from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Student, Payment
from django.db.models import Sum

def download_sales_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrolled Students"

    # Header
    headers = [
        'Salesperson',
        'Student Name',
        'Email',
        'Phone',
        'Training Class',
        'Total Fees (₹)',
        'Amount Paid (₹)',
        'Payment Status'
    ]
    for col_num, column_title in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Get data
    students = Student.objects.filter(is_enrolled=True).select_related('salesperson', 'training_class')
    for row_num, student in enumerate(students, start=2):
        total_paid = Payment.objects.filter(student=student).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0

        ws.cell(row=row_num, column=1, value=student.salesperson.user.get_full_name() if student.salesperson else "N/A")
        ws.cell(row=row_num, column=2, value=student.name)
        ws.cell(row=row_num, column=3, value=student.email)
        ws.cell(row=row_num, column=4, value=student.phone)
        ws.cell(row=row_num, column=5, value=student.training_class.name if student.training_class else "N/A")
        ws.cell(row=row_num, column=6, value=float(student.total_fees))
        ws.cell(row=row_num, column=7, value=float(total_paid))
        ws.cell(row=row_num, column=8, value=student.payment_status.capitalize())

    # Auto-resize columns
    for col_num in range(1, len(headers)+1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="enrolled_students.xlsx"'
    wb.save(response)
    return response

